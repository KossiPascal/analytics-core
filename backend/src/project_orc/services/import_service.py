"""
Service d'import OKR depuis un fichier Excel (.xlsx).

Structure attendue des feuilles :
  Colonne A : Code OKR         (ex: "Obj 1", "RC 1.1", vide pour les piliers)
  Colonne B : Priorité         ("Haute" | "Basse" | vide)
  Colonne C : Libellé ORC      (texte de l'objectif ou résultat clé)
  Colonne D : Indicateur cible (texte descriptif)
  Colonne E : Résultat final   (valeur numérique ou texte)
  Colonne F : Score            (0.00 – 1.00)
  Colonne G : Notes / fin de période

Heuristiques de détection du type de ligne :
  - Ligne vide (colonne C vide) → ignorée
  - Code commence par "Pilier" ou commence par "P" suivi d'un chiffre → PILIER
  - Code commence par "Obj" ou commence par "O" → OBJECTIF
  - Code contient "." ou commence par "RC" → RESULTAT_CLE
  - Sinon → OBJECTIF si pas de parent courant, sinon RESULTAT_CLE
"""

import re
from typing import Optional
from werkzeug.exceptions import BadRequest

from backend.src.databases.extensions import db
from backend.src.project_orc.models.pillars import StrategicPillar
from backend.src.project_orc.models.orcs    import ORC
from backend.src.project_orc.models.projects import Project
from backend.src.equipment_manager.models.employees import Department
from backend.src.logger import get_backend_logger

logger = get_backend_logger(__name__)

PRIORITY_MAP = {
    "haute":   "HIGH",
    "high":    "HIGH",
    "basse":   "LOW",
    "low":     "LOW",
    "moyenne": "MEDIUM",
    "medium":  "MEDIUM",
    "critique":"CRITICAL",
}


def _cell_str(cell) -> str:
    """Retourne la valeur d'une cellule sous forme de chaîne propre."""
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def _cell_float(cell) -> Optional[float]:
    """Retourne la valeur numérique d'une cellule ou None."""
    if cell is None or cell.value is None:
        return None
    try:
        return float(cell.value)
    except (ValueError, TypeError):
        return None


def _detect_row_type(code: str, name: str) -> str:
    """Détecte le type de ligne OKR : PILIER | OBJECTIF | RESULTAT_CLE | SKIP."""
    if not name:
        return "SKIP"
    code_upper = code.upper()
    if re.match(r"^(PILIER|P\d)", code_upper) or "PILIER" in name.upper():
        return "PILIER"
    if re.match(r"^(OBJ|OBJECTIF|O\d)", code_upper):
        return "OBJECTIF"
    if re.match(r"^(RC\s*\d|R\d|K\s*R)", code_upper) or "." in code:
        return "RESULTAT_CLE"
    # Ligne sans code mais avec texte → contexte décide
    return "UNKNOWN"


def _map_priority(raw: str) -> str:
    return PRIORITY_MAP.get(raw.lower().strip(), "MEDIUM")


def import_okr_from_xlsx(
    file_storage,
    project_id: int,
    tenant_id: int,
    user_id: int,
    fiscal_year: Optional[int],
    quarter: Optional[str],
    sheet_names: Optional[list[str]] = None,
    overwrite: bool = False,
) -> dict:
    """
    Parse un fichier Excel OKR et crée/met à jour les StrategicPillar et ORC.

    Args:
        file_storage : FileStorage Flask (request.files['file'])
        project_id   : ID du projet PROSI cible
        tenant_id    : ID du tenant courant
        user_id      : ID de l'utilisateur qui importe
        fiscal_year  : Année fiscale (ex: 2026)
        quarter      : Trimestre cible (T1 | T2 | T3 | T4 | YEARLY)
        sheet_names  : Liste des feuilles à importer (None = toutes)
        overwrite    : Si True, met à jour les ORCs existants par (project_id, code, fiscal_year, quarter)

    Returns:
        dict avec les compteurs : pillars_created, orcs_created, orcs_updated, errors
    """
    import openpyxl

    # ── Vérifier le projet ─────────────────────────────────────────────────
    project = Project.query.filter_by(id=project_id, tenant_id=tenant_id, deleted=False).first()
    if not project:
        raise BadRequest("Projet introuvable")

    try:
        wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
    except Exception as e:
        raise BadRequest(f"Fichier Excel invalide : {e}")

    # Sélectionner les feuilles à traiter
    sheets_to_process = sheet_names if sheet_names else wb.sheetnames

    stats = {
        "pillars_created": 0,
        "orcs_created":    0,
        "orcs_updated":    0,
        "skipped":         0,
        "errors":          [],
    }

    # Index des départements par nom (case-insensitive)
    dept_index: dict[str, int] = {}
    for dept in Department.query.filter_by(deleted=False).all():
        dept_index[dept.name.lower().strip()] = dept.id

    for sheet_name in sheets_to_process:
        if sheet_name not in wb.sheetnames:
            stats["errors"].append(f"Feuille '{sheet_name}' introuvable")
            continue

        # Résoudre le département via le nom de la feuille
        dept_id = dept_index.get(sheet_name.lower().strip())

        # Chercher ou créer un pilier "par défaut" pour cette feuille si aucun pilier n'est détecté
        ws = wb[sheet_name]
        _import_sheet(
            ws, sheet_name, project, tenant_id, user_id,
            fiscal_year, quarter, dept_id, overwrite, stats
        )

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error(f"Erreur commit import OKR: {e}")
        raise BadRequest(f"Erreur lors de la sauvegarde : {e}")

    wb.close()
    return stats


def _import_sheet(ws, sheet_name, project, tenant_id, user_id,
                  fiscal_year, quarter, dept_id, overwrite, stats):
    """Traite une feuille Excel et crée les entités correspondantes."""

    current_pillar: Optional[StrategicPillar] = None
    current_objectif: Optional[ORC] = None
    pillar_order = 0

    for row in ws.iter_rows(min_row=2):  # Ligne 1 = en-tête
        if len(row) < 3:
            continue

        code  = _cell_str(row[0])
        raw_priority = _cell_str(row[1]) if len(row) > 1 else ""
        name  = _cell_str(row[2])
        target_indicator = _cell_str(row[3]) if len(row) > 3 else ""
        result_raw       = _cell_str(row[4]) if len(row) > 4 else ""
        score_raw        = _cell_float(row[5]) if len(row) > 5 else None
        notes            = _cell_str(row[6]) if len(row) > 6 else ""

        row_type = _detect_row_type(code, name)

        if row_type == "SKIP":
            continue

        priority = _map_priority(raw_priority) if raw_priority else "MEDIUM"

        # Résultat final → current_value
        current_value = None
        try:
            current_value = float(result_raw) if result_raw else None
        except ValueError:
            pass

        # ── PILIER ────────────────────────────────────────────────────────
        if row_type == "PILIER":
            pillar_order += 1
            pillar_code = code or f"P{pillar_order}"
            current_pillar = StrategicPillar.query.filter_by(
                project_id=project.id, code=pillar_code, deleted=False
            ).first()
            if not current_pillar:
                current_pillar = StrategicPillar(
                    tenant_id=tenant_id,
                    project_id=project.id,
                    name=name,
                    code=pillar_code,
                    order_index=pillar_order,
                    fiscal_year=fiscal_year,
                    created_by=user_id,
                    updated_by=user_id,
                )
                db.session.add(current_pillar)
                db.session.flush()  # Pour obtenir l'id
                stats["pillars_created"] += 1
            current_objectif = None
            continue

        # ── OBJECTIF ──────────────────────────────────────────────────────
        if row_type == "OBJECTIF" or (row_type == "UNKNOWN" and current_objectif is None):
            orc = _find_or_create_orc(
                code=code, name=name, project=project, tenant_id=tenant_id,
                user_id=user_id, orc_type="OBJECTIF",
                pillar_id=current_pillar.id if current_pillar else None,
                parent_id=None, dept_id=dept_id,
                priority=priority, target_indicator=target_indicator,
                current_value=current_value, score=score_raw, notes=notes,
                fiscal_year=fiscal_year, quarter=quarter,
                overwrite=overwrite, stats=stats,
            )
            current_objectif = orc
            continue

        # ── RESULTAT_CLE ──────────────────────────────────────────────────
        if row_type == "RESULTAT_CLE" or (row_type == "UNKNOWN" and current_objectif is not None):
            _find_or_create_orc(
                code=code, name=name, project=project, tenant_id=tenant_id,
                user_id=user_id, orc_type="RESULTAT_CLE",
                pillar_id=current_pillar.id if current_pillar else None,
                parent_id=current_objectif.id if current_objectif else None,
                dept_id=dept_id,
                priority=priority, target_indicator=target_indicator,
                current_value=current_value, score=score_raw, notes=notes,
                fiscal_year=fiscal_year, quarter=quarter,
                overwrite=overwrite, stats=stats,
            )


def _find_or_create_orc(
    code, name, project, tenant_id, user_id, orc_type,
    pillar_id, parent_id, dept_id, priority, target_indicator,
    current_value, score, notes, fiscal_year, quarter, overwrite, stats
) -> ORC:
    """Cherche un ORC existant par (project_id, code, fiscal_year, quarter) ou le crée."""

    existing = None
    if code and overwrite:
        existing = ORC.query.filter_by(
            project_id=project.id,
            code=code,
            fiscal_year=fiscal_year,
            quarter=quarter,
            deleted=False,
        ).first()

    if existing:
        existing.name             = name
        existing.priority         = priority
        existing.target_indicator = target_indicator
        existing.current_value    = current_value
        existing.score            = score
        existing.notes            = notes
        existing.pillar_id        = pillar_id
        existing.department_id    = dept_id
        existing.updated_by       = user_id
        stats["orcs_updated"] += 1
        return existing
    else:
        orc = ORC(
            tenant_id=tenant_id,
            project_id=project.id,
            pillar_id=pillar_id,
            parent_id=parent_id,
            department_id=dept_id,
            orc_type=orc_type,
            code=code,
            name=name,
            priority=priority,
            target_indicator=target_indicator,
            current_value=current_value,
            score=score,
            notes=notes,
            fiscal_year=fiscal_year,
            quarter=quarter,
            status="DRAFT",
            created_by=user_id,
            updated_by=user_id,
        )
        db.session.add(orc)
        db.session.flush()
        stats["orcs_created"] += 1
        return orc
